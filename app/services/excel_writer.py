from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional, Set

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.product_schema import ProductItem
from app.schemas.sales_schema import SalesLineItem


EXCEL_COLUMNS = [
    "商品名称",
    "商品单价",
    "购买份数",
    "单份数量",
    "量词",
    "总数量",
    "商品金额",
    "消耗类型",
    "订单生成时间",
    "备注",
]

SALES_EXCEL_COLUMNS = [
    "项目/菜品名称",
    "数量",
    "单价",
    "金额",
    "识别时间",
    "备注",
]


def _estimate_display_width(value: object) -> int:
    text = "" if value is None else str(value)
    # Chinese chars are generally wider than ASCII in Excel.
    width = 0
    for ch in text:
        width += 2 if ord(ch) > 127 else 1
    return max(width, 4)


def _beautify_sheet(ws, *, max_col: int, wide_text_cols: Optional[Set[int]] = None) -> None:
    wide_text_cols = wide_text_cols or set()
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    header_font = Font(bold=True)
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    max_widths: dict[int, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            max_widths[cell.column] = max(max_widths.get(cell.column, 0), _estimate_display_width(cell.value))

    for col_idx in range(1, max_col + 1):
        col_name = get_column_letter(col_idx)
        base = max(max_widths.get(col_idx, 10), 12)
        if col_idx in wide_text_cols:
            base = max(base, 26)
        ws.column_dimensions[col_name].width = min(base, 60)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            text = "" if cell.value is None else str(cell.value)
            if cell.column in wide_text_cols or len(text) > 26:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center")


def write_items_to_excel(items: Iterable[ProductItem], output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "商品明细"
    ws.append(EXCEL_COLUMNS)

    for item in items:
        raw_type = getattr(item, "consumption_type", None)
        type_label = ""
        if raw_type == "instant":
            type_label = "即时消耗"
        elif raw_type == "non_instant":
            type_label = "非即时消耗"
        elif raw_type:
            type_label = str(raw_type)
        row = [
            item.product_name,
            item.unit_price,
            item.quantity,
            item.unit_amount,
            item.quantity_unit,
            item.total_quantity,
            item.amount,
            type_label,
            item.order_created_at.strftime("%Y-%m-%d %H:%M:%S"),
            item.remarks,
        ]
        ws.append(row)

    _beautify_sheet(ws, max_col=len(EXCEL_COLUMNS), wide_text_cols={1, len(EXCEL_COLUMNS)})

    filename = f"商品报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = output_dir / filename
    wb.save(file_path)
    return file_path


def write_sales_to_excel(items: Iterable[SalesLineItem], output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "营业明细"
    ws.append(SALES_EXCEL_COLUMNS)

    for item in items:
        ws.append(
            [
                item.item_name,
                item.quantity,
                item.unit_price,
                item.amount,
                item.order_created_at.strftime("%Y-%m-%d %H:%M:%S"),
                item.remarks,
            ]
        )

    _beautify_sheet(ws, max_col=len(SALES_EXCEL_COLUMNS), wide_text_cols={1, len(SALES_EXCEL_COLUMNS)})

    filename = f"营业报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = output_dir / filename
    wb.save(file_path)
    return file_path


def write_daily_report_excel(report: dict, output_dir: Path) -> Path:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "日报汇总"
    ws_summary.append(["字段", "数值"])
    ws_summary.append(["营业日期", report.get("biz_date", "")])
    ws_summary.append(["营业收入", report.get("营业收入", report.get("revenue_total", 0))])
    ws_summary.append(["采购总额", report.get("采购总额", report.get("purchase_total", 0))])
    ws_summary.append(["即时消耗成本估算", report.get("即时消耗成本估算", report.get("consumed_cost_estimate", 0))])
    ws_summary.append(["非即时耗材费用", report.get("非即时耗材费用", report.get("non_instant_expense_total", 0))])
    ws_summary.append(["毛利估算", report.get("毛利估算", report.get("gross_profit", 0))])
    _beautify_sheet(ws_summary, max_col=2, wide_text_cols={1, 2})

    def _write_sheet(title: str, rows: list[dict]) -> None:
        ws = wb.create_sheet(title=title)
        if not rows:
            ws.append(["暂无数据"])
            _beautify_sheet(ws, max_col=1, wide_text_cols={1})
            return
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(k, "") for k in headers])
        wide_cols = {idx + 1 for idx, h in enumerate(headers) if any(x in h for x in ("备注", "名称", "说明"))}
        _beautify_sheet(ws, max_col=len(headers), wide_text_cols=wide_cols)

    _write_sheet("菜品销售", report.get("dish_sales", []))
    _write_sheet("用料消耗", report.get("ingredient_consumption", []))
    _write_sheet("库存剩余", report.get("inventory_remaining", []))
    _write_sheet("菜品耗料明细", report.get("dish_inventory_consumption", []))
    _write_sheet("缺料预警", report.get("shortage_alerts", []))
    _write_sheet("异常明细", report.get("exception_details", []))

    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"日报_{report.get('biz_date', datetime.now().date().isoformat())}_{datetime.now().strftime('%H%M%S')}.xlsx"
    path = output_dir / filename
    wb.save(path)
    return path
